import os
import sys
import platform
import datetime
import webbrowser
import logging
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, filedialog
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment
from PIL import ImageTk, Image

logging.disable(logging.DEBUG)
logging.basicConfig(filename='class_management.log', level=logging.DEBUG,
                    format='%(funcName)s - %(asctime)s: %(message)s',
                    datefmt='%Y-%m-%d %H:%M')

school_color_1 = 'red'
school_color_2 = '#484848'
school_color_3 = 'white'


def open_url():
    logging.debug('Opening help page...')
    webbrowser.open('https://github.com/thomaskellough/Personal-Projects/blob/class_management/class_management_log/how-to.md')


def display_license():
    tk.messagebox.showinfo(
        message='The MIT License (MIT)\n'
                '=====================\n\n'
                'Copyright (c) 2018 Thomas Kellough\n\n'
                'Permission is hereby granted, free of charge, to any person obtaining a copy of\n'
                'this software and associated documentation files (the "Software"), to deal in\n'
                'the Software without restriction, including without limitation the rights to\n'
                'use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies\n'
                'of the Software, and to permit persons to whom the Software is furnished to do\n'
                'so, subject to the following conditions:\n\n'

                'The above copyright notice and this permission notice shall be included in all'
                'copies or substantial portions of the Software.\n\n'

                'THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR\n'
                'IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,\n'
                'FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE\n'
                'AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER\n'
                'LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,\n'
                'OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE\n'
                'SOFTWARE.')


def write_excel_page(wb, name):
    logging.debug('Writing excel page...')
    name = name.replace('\n', '')
    wb.create_sheet(name)
    active_sheet = wb[name]
    active_sheet['A1'] = 'Date'
    active_sheet['B1'] = 'Infraction'
    active_sheet['C1'] = 'Consequence Level'
    active_sheet['D1'] = 'Notes'
    active_sheet['A1'].style = 'Accent1'
    active_sheet['B1'].style = 'Accent2'
    active_sheet['C1'].style = 'Accent6'
    active_sheet['D1'].style = 'Accent4'
    active_sheet.column_dimensions['A'].width = 20
    active_sheet.column_dimensions['B'].width = 30
    active_sheet.column_dimensions['C'].width = 30
    active_sheet.column_dimensions['D'].width = 30


def open_file():
    var = tk.filedialog.askopenfilename(initialdir='.', title='Select File', filetypes=(('Excel files', ('*.xl*', '*.xlsx', '*.xlsm', '*.xlsb', '.xlam', '*.xltx', '*.xltm', '*.xls', '*.xla', '*.xlt', '*.xlm', '*.xlw')), ('Text Files', ('*.txt', '*.csv')), ('all files', '*.*')))
    os.startfile(var)


def get_student_names(workbook):
    try:
        wb = load_workbook(workbook)
        names = wb.sheetnames
        logging.debug(f'Returning: {names}')
        return names
    except FileNotFoundError:
        logging.debug(f'{workbook} not found.')


# Obtain student names from each workbook
class_list = ['period_1.xlsx', 'period_2.xlsx', 'period_3.xlsx', 'period_4.xlsx',
              'period_5.xlsx', 'period_6.xlsx', 'period_7.xlsx', 'period_8.xlsx', ]

infractions = [
    'A: Behavior Contract',
    'B: Not prepared for class',
    'C: Refusal to follow instructions',
    'D: Distracting behavior during class',
    'E: Inappropriate use of technology',
    'F: Threatening behaviour to others',
    'G: Not attending tutoring for missed assignments',
    'H: Failure to participate in class',
    'I: Sleeping during class',
    'J: Inappropriate comments'
]
consequences = [
    '1) Verbal Warning',
    '2) Student/Teacher Conference',
    '3) Administrative Communications',
    '4) Parent Contact & Detention',
    '5) Office Referral'
]


def set_icon(app):
    if platform.system() == "Windows":
        app.iconbitmap(resource_path(os.path.join("Assets", "icon.ico")))
    elif platform.system() == "Linux":
        imgicon = tk.PhotoImage(file=resource_path(os.path.join("Assets", "icon.gif")))
        app.tk.call("wm", "iconphoto", app._w, imgicon)


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class App(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        root.configure(bg=school_color_2, padx=5)
        root.title('Classroom Management Log')
        toolbar = tk.Frame(master, bg=school_color_2, bd=1, relief=tk.RAISED)
        upper_frame = tk.Frame(master, bg=school_color_2)
        lower_frame = tk.Frame(master, bg=school_color_2)
        toolbar.grid(column=0, row=0, sticky=tk.EW)
        upper_frame.grid(column=0, row=1)
        lower_frame.grid(column=0, row=2, padx=10, pady=5)

        self.roster_1 = get_student_names(class_list[0])
        self.roster_2 = get_student_names(class_list[1])
        self.roster_3 = get_student_names(class_list[2])
        self.roster_4 = get_student_names(class_list[3])
        self.roster_5 = get_student_names(class_list[4])
        self.roster_6 = get_student_names(class_list[5])
        self.roster_7 = get_student_names(class_list[6])
        self.roster_8 = get_student_names(class_list[7])

        self.roster_dict = {'Period 1': self.roster_1, 'Period 2': self.roster_2, 'Period 3': self.roster_3,
                            'Period 4': self.roster_4, 'Period 5': self.roster_5, 'Period 6': self.roster_6,
                            'Period 7': self.roster_7, 'Period 8': self.roster_8}
        self.roster_dict = dict((k, v) for k, v in self.roster_dict.items() if v is not None)

        self.var_to_excel_dict = {'Period 1': 'period_1.xlsx', 'Period 2': 'period_2.xlsx', 'Period 3': 'period_3.xlsx',
                                  'Period 4': 'period_4.xlsx', 'Period 5': 'period_5.xlsx', 'Period 6': 'period_6.xlsx',
                                  'Period 7': 'period_7.xlsx', 'Period 8': 'period_8.xlsx'}

        # Set variables
        self.period_select_var = tk.StringVar(self)
        self.student_select_var = tk.StringVar(self)
        self.infraction_select_var = tk.StringVar(self)
        self.consequence_level_var = tk.StringVar(self)
        self.notes_text_var = tk.StringVar(self)
        self.add_student_var = tk.StringVar(self)
        self.transfer_from_var = tk.StringVar(self)
        self.transfer_to_var = tk.StringVar(self)
        self.student_transfer_var = tk.StringVar(self)
        self.period_select_var.trace('w', self.update_student_options)
        self.transfer_from_var.trace('w', self.update_student_options_transfer)

        # Toolbar
        relief = tk.FLAT
        refresh_img = ImageTk.PhotoImage(Image.open(resource_path('Assets\\refreshicon.png')))
        self.refresh_ui_btn = tk.Button(toolbar, image=refresh_img, relief=relief, command=self.refresh_ui)
        self.refresh_ui_btn.pack(side=tk.LEFT, padx=1)
        self.refresh_ui_btn.bind('<Enter>', self.refresh_enter)
        self.refresh_ui_btn.bind('<Leave>', self.close)
        self.refresh_tooltip = 'Refresh window...'

        open_img = ImageTk.PhotoImage(Image.open(resource_path('Assets\openicon.png')))
        self.open_btn = tk.Button(toolbar, image=open_img, relief=relief, command=open_file)
        self.open_btn.pack(side=tk.LEFT, padx=1)
        self.open_btn.bind('<Enter>', self.open_enter)
        self.open_btn.bind('<Leave>', self.close)
        self.open_tooltip = 'Open...'

        save_img = ImageTk.PhotoImage(Image.open(resource_path('Assets\saveicon.png')))
        self.save_btn = tk.Button(toolbar, image=save_img, relief=relief, command=self.save_file)
        self.save_btn.pack(side=tk.LEFT, padx=1)
        self.save_btn.bind('<Enter>', self.save_enter)
        self.save_btn.bind('<Leave>', self.close)
        self.save_tooltip = 'Save...'

        add_img = ImageTk.PhotoImage(Image.open(resource_path('Assets\\addicon.png')))
        self.add_student_btn = tk.Button(toolbar, image=add_img, relief=relief, command=self.add_student_window)
        self.add_student_btn.pack(side=tk.LEFT, padx=1)
        self.add_student_btn.bind('<Enter>', self.add_enter)
        self.add_student_btn.bind('<Leave>', self.close)
        self.add_tooltip = 'Add a student...'

        delete_img = ImageTk.PhotoImage(Image.open(resource_path('Assets\deleteicon.png')))
        self.delete_student_btn = tk.Button(toolbar, image=delete_img, relief=relief, command=self.del_student)
        self.delete_student_btn.pack(side=tk.LEFT, padx=1)
        self.delete_student_btn.bind('<Enter>', self.del_enter)
        self.delete_student_btn.bind('<Leave>', self.close)
        self.del_tooltip = 'Delete student(s)...'

        transfer_img = ImageTk.PhotoImage(Image.open(resource_path('Assets\\transfericon.png')))
        self.transfer_student_btn = tk.Button(toolbar, image=transfer_img, relief=relief, command=self.transfer_student)
        self.transfer_student_btn.pack(side=tk.LEFT, padx=1)
        self.transfer_student_btn.bind('<Enter>', self.transfer_enter)
        self.transfer_student_btn.bind('<Leave>', self.close)
        self.transfer_tooltip = 'Transfer student...'

        # Main Window Logo
        logo_img = ImageTk.PhotoImage(Image.open(resource_path('Assets\logo.png')))
        self.logo_img_label = tk.Label(upper_frame, image=logo_img, bg='lightgrey')

        # Widgets
        self.title_text = ttk.Label(upper_frame, text=' Class Management Log ', style='title.TLabel', borderwidth=2,
                                    relief=tk.GROOVE, background='lightgrey')
        self.period_select_menu = ttk.OptionMenu(upper_frame, self.period_select_var,
                                                 'Period 1', *self.roster_dict.keys(),
                                                 style='op_menu.TMenubutton')
        self.student_select_menu = ttk.OptionMenu(upper_frame,
                                                  self.student_select_var, '',
                                                  style='op_menu.TMenubutton')
        self.infraction_select_menu = ttk.OptionMenu(lower_frame,
                                                     self.infraction_select_var,
                                                     infractions[0],
                                                     *infractions,
                                                     style='op_menu.TMenubutton')
        self.consequence_level_menu = ttk.OptionMenu(lower_frame,
                                                     self.consequence_level_var,
                                                     consequences[0],
                                                     *consequences,
                                                     style='op_menu.TMenubutton')
        self.notes_label = ttk.Label(lower_frame,
                                     text='Notes:',
                                     style='notes.TLabel')
        self.notes_entry = ttk.Entry(lower_frame,
                                     textvariable=self.notes_text_var,
                                     width=80)
        self.submit_btn = ttk.Button(lower_frame,
                                     text='Submit',
                                     command=self.write_to_excel)
        self.period_select_var.set('Period 1')

        # Styling
        s = ttk.Style()
        s.configure('op_menu.TMenubutton', font='Tahoma 14', foreground=school_color_3, background=school_color_2)
        s.configure('notes.TLabel', font='Tahoma 12', foreground=school_color_3, background=school_color_2)
        s.configure('title.TLabel', font='Tahoma 28', foreground=school_color_1, background=school_color_2)

        # Packing
        self.logo_img_label.grid(column=0, row=0, sticky=tk.NSEW)
        self.title_text.grid(column=1, row=0, sticky=tk.NSEW)
        self.period_select_menu.grid(column=0, row=1, sticky=tk.EW, columnspan=2)
        self.student_select_menu.grid(column=0, row=2, sticky=tk.EW, columnspan=2)
        self.infraction_select_menu.grid(sticky=tk.W)
        self.consequence_level_menu.grid(sticky=tk.W)
        self.notes_label.grid(sticky=tk.W)
        self.notes_entry.grid(sticky=tk.W)
        self.submit_btn.grid(sticky=tk.W)

        # Menus
        root.option_add('*tearOff', tk.FALSE)
        main_menu = tk.Menu(root)
        root.configure(menu=main_menu)
        sub_menu_file = tk.Menu(lower_frame)
        sub_menu_edit = tk.Menu(main_menu)
        sub_menu_help = tk.Menu(main_menu)

        main_menu.add_cascade(label='File', menu=sub_menu_file)
        sub_menu_file.add_command(label='Open...', command=open_file)
        sub_menu_file.add_command(label='Save...', command=self.save_file)
        sub_menu_file.add_separator()
        sub_menu_file.add_command(label='Create Excel Files...', command=self.create_excel_files)
        sub_menu_file.add_separator()
        sub_menu_file.add_command(label='Exit', command=root.quit)

        main_menu.add_cascade(label='Edit', menu=sub_menu_edit)
        sub_menu_edit.add_command(label='Add student...', command=self.add_student_window)
        sub_menu_edit.add_command(label='Delete student(s)...', command=self.del_student)
        sub_menu_edit.add_command(label='Transfer student...', command=self.transfer_student)

        main_menu.add_cascade(label='Help', menu=sub_menu_help)
        sub_menu_help.add_command(label='How to use this program...', command=open_url)
        sub_menu_help.add_separator()
        sub_menu_help.add_command(label='License', command=display_license)

    def update_student_options(self, *args):
        students = self.roster_dict[self.period_select_var.get()]
        self.student_select_var.set(students[0])

        menu = self.student_select_menu['menu']
        menu.delete(0, 'end')

        for student in students:
            menu.add_command(label=student, command=lambda period=student: self.student_select_var.set(period))

    def update_student_options_transfer(self, *args):
        students = self.roster_dict[self.transfer_from_var.get()]
        self.student_transfer_var.set(students[0])

        menu = self.student_transfer_menu['menu']
        menu.delete(0, 'end')

        for student in students:
            menu.add_command(label=student, command=lambda period=student: self.student_transfer_var.set(period))

    # Toolbar tooltips
    def refresh_ui(self):
        logging.debug('Refreshing UI...')
        self.roster_1 = get_student_names(class_list[0])
        self.roster_2 = get_student_names(class_list[1])
        self.roster_3 = get_student_names(class_list[2])
        self.roster_4 = get_student_names(class_list[3])
        self.roster_5 = get_student_names(class_list[4])
        self.roster_6 = get_student_names(class_list[5])
        self.roster_7 = get_student_names(class_list[6])
        self.roster_8 = get_student_names(class_list[7])

        self.roster_dict = {'Period 1': self.roster_1, 'Period 2': self.roster_2, 'Period 3': self.roster_3,
                            'Period 4': self.roster_4, 'Period 5': self.roster_5, 'Period 6': self.roster_6,
                            'Period 7': self.roster_7, 'Period 8': self.roster_8}
        self.roster_dict = dict((k, v) for k, v in self.roster_dict.items() if v is not None)
        self.update_student_options()

    def refresh_enter(self, event=None):
        x = y = 0
        x, y, cx, cy = self.refresh_ui_btn.bbox('insert')
        x += self.refresh_ui_btn.winfo_rootx() + 25
        y += self.refresh_ui_btn.winfo_rooty() + 20
        # Creates a toplevel window
        self.tw = tk.Toplevel(self.refresh_ui_btn)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry('+%d+%d' % (x, y))
        label = tk.Label(self.tw, text=self.refresh_tooltip, justify='left',
                         background='yellow', relief='solid', borderwidth=1,
                         font=('times', '10', 'normal'))
        label.pack(ipadx=1)

    def save_enter(self, event=None):
        x = y = 0
        x, y, cx, cy = self.save_btn.bbox('insert')
        x += self.save_btn.winfo_rootx() + 25
        y += self.save_btn.winfo_rooty() + 20
        # Creates a toplevel window
        self.tw = tk.Toplevel(self.save_btn)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry('+%d+%d' % (x, y))
        label = tk.Label(self.tw, text=self.save_tooltip, justify='left',
                         background='yellow', relief='solid', borderwidth=1,
                         font=('times', '10', 'normal'))
        label.pack(ipadx=1)

    def open_enter(self, event=None):
        x = y = 0
        x, y, cx, cy = self.open_btn.bbox('insert')
        x += self.open_btn.winfo_rootx() + 25
        y += self.open_btn.winfo_rooty() + 20
        # Creates a toplevel window
        self.tw = tk.Toplevel(self.open_btn)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry('+%d+%d' % (x, y))
        label = tk.Label(self.tw, text=self.open_tooltip, justify='left',
                         background='yellow', relief='solid', borderwidth=1,
                         font=('times', '10', 'normal'))
        label.pack(ipadx=1)

    def add_enter(self, event=None):
        x = y = 0
        x, y, cx, cy = self.add_student_btn.bbox('insert')
        x += self.add_student_btn.winfo_rootx() + 25
        y += self.add_student_btn.winfo_rooty() + 20
        # Creates a toplevel window
        self.tw = tk.Toplevel(self.add_student_btn)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry('+%d+%d' % (x, y))
        label = tk.Label(self.tw, text=self.add_tooltip, justify='left',
                         background='yellow', relief='solid', borderwidth=1,
                         font=('times', '10', 'normal'))
        label.pack(ipadx=1)

    def del_enter(self, event=None):
        x = y = 0
        x, y, cx, cy = self.delete_student_btn.bbox('insert')
        x += self.delete_student_btn.winfo_rootx() + 25
        y += self.delete_student_btn.winfo_rooty() + 20
        # Creates a toplevel window
        self.tw = tk.Toplevel(self.delete_student_btn)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry('+%d+%d' % (x, y))
        label = tk.Label(self.tw, text=self.del_tooltip, justify='left',
                         background='yellow', relief='solid', borderwidth=1,
                         font=('times', '10', 'normal'))
        label.pack(ipadx=1)

    def transfer_enter(self, event=None):
        x = y = 0
        x, y, cx, cy = self.transfer_student_btn.bbox('insert')
        x += self.transfer_student_btn.winfo_rootx() + 25
        y += self.transfer_student_btn.winfo_rooty() + 20
        # Creates a toplevel window
        self.tw = tk.Toplevel(self.transfer_student_btn)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry('+%d+%d' % (x, y))
        label = tk.Label(self.tw, text=self.transfer_tooltip, justify='left',
                         background='yellow', relief='solid', borderwidth=1,
                         font=('times', '10', 'normal'))
        label.pack(ipadx=1)

    def close(self, event=None):
        if self.tw:
            self.tw.destroy()

    # Toolbar functions
    # open_file is static
    def save_file(self):
        var = tk.filedialog.asksaveasfilename(initialdir='.', title='Save file', filetypes=(('Excel files', ('*.xl*', '*.xlsx', '*.xlsm', '*.xlsb', '.xlam', '*.xltx', '*.xltm', '*.xls', '*.xla', '*.xlt', '*.xlm', '*.xlw')), ('Text Files', ('*.txt', '*.csv')), ('all files', '*.*')))
        wb = load_workbook(self.var_to_excel_dict[self.period_select_var.get()])
        sheets = wb.sheetnames
        for sheet in sheets:
            if sheet != self.student_select_var.get():
                del wb[sheet]
        wb.save(var + '.xlsx')

    def add_student_window(self):
        def add_student():
            wb = load_workbook(self.var_to_excel_dict[self.period_select_var.get()])
            name = self.add_student_var.get()
            answer = tk.messagebox.askquestion(title='Adding student...',
                                               message=f'Are you sure you want to add {name} to '
                                                       f'{self.period_select_var.get()}?')
            if answer == 'yes':
                logging.debug('Adding a new student...')
                write_excel_page(wb, name)
                wb.save(self.var_to_excel_dict[self.period_select_var.get()])
                tk.messagebox.showinfo(title='Update Message', message='Student added successfully.')
                self.refresh_ui()
                logging.debug('New student added.')
                add_student_window.destroy()
        add_student_window = tk.Toplevel(root)
        set_icon(add_student_window)
        add_student_window.configure(bg=school_color_2)
        add_student_window.title('Add a student')
        # Widgets
        add_student_text = ttk.Label(add_student_window, text='Please type in a name of a student to add.',
                                     style='notes.TLabel')
        add_student_entry = ttk.Entry(add_student_window, textvariable=self.add_student_var)
        add_student_btn = ttk.Button(add_student_window, text='Add Student', command=add_student)
        # Packing
        add_student_text.grid(column=0, row=0, sticky='w', padx=5, pady=5)
        add_student_entry.grid(column=0, row=1, sticky='w', padx=5, pady=5)
        add_student_btn.grid(column=0, row=2, sticky='w', padx=5, pady=5)

    def del_student(self):
        def delete_sheets():
            name_dict = {}
            remove_dict = {}
            for i, student in enumerate(self.roster_dict[self.period_select_var.get()]):
                name_dict[i] = student
            for j in remove_student_listbox.curselection():
                remove_dict[j] = name_dict[j]
            remove_name_list = []
            for name in remove_dict.values():
                remove_name_list.append(name)
            answer = tk.messagebox.askquestion(title='Removing students...',
                                               message=f'Are you sure you want to remove {",".join(remove_name_list)} '
                                                       f'from {self.period_select_var.get()}?\nThis will delete all '
                                                       f'data already documented.')
            if answer == 'yes':
                logging.debug('Removing a current student...')
                wb = load_workbook(self.var_to_excel_dict[self.period_select_var.get()])
                for name in remove_name_list:
                    wb.remove(wb[name])
                wb.save(self.var_to_excel_dict[self.period_select_var.get()])
                tk.messagebox.showinfo(title='Update Message', message='Students deleted successfully.')
                self.refresh_ui()
                logging.debug('Student removed.')
                remove_student_window.destroy()

        remove_student_window = tk.Toplevel(root)
        set_icon(remove_student_window)
        remove_student_window.configure(bg=school_color_2)
        remove_student_window.title('Remove student(s)')
        remove_student_listbox = tk.Listbox(remove_student_window, selectmode=tk.MULTIPLE)
        for name in self.roster_dict[self.period_select_var.get()]:
            remove_student_listbox.insert(tk.END, name)
        remove_student_btn = ttk.Button(remove_student_window, text='Remove', command=delete_sheets)
        remove_student_listbox.grid(padx=5, pady=5)
        remove_student_btn.grid(padx=5, pady=5)

    def transfer_student_submit(self):
        try:
            wb_from = load_workbook(self.var_to_excel_dict[self.transfer_from_var.get()])
            ws_from = wb_from[self.student_transfer_var.get()]
            wb_to = load_workbook(self.var_to_excel_dict[self.transfer_to_var.get()])
            ws_to = wb_to.create_sheet(self.student_transfer_var.get())
            max_row = ws_from.max_row
            for column in range(1, 5):
                for row in range(1, max_row + 1):
                    ws_to.cell(row=row, column=column).value = ws_from.cell(row=row, column=column).value
                    ws_to.cell(row=max_row, column=column).alignment = Alignment(vertical='top', wrapText=True)
            # Style copied cells
            ws_to['A1'].style = 'Accent1'
            ws_to['B1'].style = 'Accent2'
            ws_to['C1'].style = 'Accent6'
            ws_to['D1'].style = 'Accent4'
            ws_to.column_dimensions['A'].width = 20
            ws_to.column_dimensions['B'].width = 30
            ws_to.column_dimensions['C'].width = 30
            ws_to.column_dimensions['D'].width = 30
            for row in range(2, ws_to.max_row + 1):
                ws_to.cell(row=row, column=1).fill = PatternFill(start_color='b7dee8',
                                                                 end_color='b7dee8',
                                                                 fill_type='solid')
            for row in range(2, ws_to.max_row + 1):
                ws_to.cell(row=row, column=2).fill = PatternFill(start_color='e6b8b7',
                                                                 end_color='e6b8b7',
                                                                 fill_type='solid')
            for row in range(2, ws_to.max_row + 1):
                ws_to.cell(row=row, column=3).fill = PatternFill(start_color='fcd5b4',
                                                                 end_color='fcd5b4',
                                                                 fill_type='solid')
            for row in range(2, ws_to.max_row + 1):
                ws_to.cell(row=row, column=4).fill = PatternFill(start_color='ccc0da',
                                                                 end_color='ccc0da',
                                                                 fill_type='solid')
            wb_from.remove(wb_from[self.student_transfer_var.get()])
            wb_to.save(self.var_to_excel_dict[self.transfer_to_var.get()])
            wb_from.save(self.var_to_excel_dict[self.transfer_from_var.get()])
            wb_from.close()
            wb_to.close()
            tk.messagebox.showinfo('Transfer successful', message=f'{self.student_transfer_var.get()} was successfully '
                                   f'transferred from {self.transfer_from_var.get()} to {self.transfer_to_var.get()}.')
            self.refresh_ui()
            self.new_window.destroy()
        except KeyError:
            logging.debug('Error occurred. Could not transfer student.')
            tk.messagebox.showerror('Error occurred!', message='An error has occurred. Please try restarting the'
                                    'program.')

    def transfer_student(self):
        self.new_window = tk.Toplevel(root)
        set_icon(self.new_window)
        self.new_window.configure(bg=school_color_2)
        self.new_window.grid()
        self.student_transfer_menu = ttk.OptionMenu(self.new_window,
                                                    self.student_transfer_var, '',
                                                    style='op_menu.TMenubutton')
        # Widgets
        transfer_from_options = ttk.OptionMenu(self.new_window, self.transfer_from_var,
                                               'Period 1', *self.roster_dict.keys(),
                                               style='op_menu.TMenubutton')
        transfer_to_options = ttk.OptionMenu(self.new_window, self.transfer_to_var,
                                             'Period 1', *self.roster_dict.keys(),
                                             style='op_menu.TMenubutton')
        transfer_btn = ttk.Button(self.new_window, text='Submit', command=self.transfer_student_submit)
        transfer_from_options.grid(column=0, row=0)
        transfer_to_options.grid(column=2, row=0)
        self.student_transfer_menu.grid(column=1, row=0)
        transfer_btn.grid(column=0, row=1)

    def write_to_excel(self, var1=None, var2=None):
        logging.debug('Writing a new line to Excel.')
        if var1 is None:
            var1 = self.period_select_var.get()
        if var2 is None:
            var2 = self.student_select_var.get()
        wb = load_workbook(self.var_to_excel_dict[var1])
        active_sheet = wb[var2]
        row = active_sheet.max_row + 1
        active_sheet.cell(row=row, column=1).value = datetime.datetime.now()
        active_sheet.cell(row=row, column=1).fill = PatternFill(start_color='b7dee8',
                                                                end_color='b7dee8',
                                                                fill_type='solid')
        active_sheet.cell(row=row, column=2).value = self.infraction_select_var.get()
        active_sheet.cell(row=row, column=2).fill = PatternFill(start_color='e6b8b7',
                                                                end_color='e6b8b7',
                                                                fill_type='solid')
        active_sheet.cell(row=row, column=3).value = self.consequence_level_var.get()
        active_sheet.cell(row=row, column=3).fill = PatternFill(start_color='fcd5b4',
                                                                end_color='fcd5b4',
                                                                fill_type='solid')
        active_sheet.cell(row=row, column=4).value = self.notes_text_var.get()
        active_sheet.cell(row=row, column=4).fill = PatternFill(start_color='ccc0da',
                                                                end_color='ccc0da',
                                                                fill_type='solid')
        for i in range(1, 5):
            active_sheet.cell(row=row, column=i).alignment = Alignment(vertical='top', wrapText=True)
        wb.save(self.var_to_excel_dict[var1])
        wb.close()
        logging.debug('Message recorded and saved.')
        tk.messagebox.showinfo('Verification', message='Message logged.')

    def create_excel_files(self):
        answer = tk.messagebox.askquestion(title='Creating files...',
                                           message=f'Clicking yes will enable new excel files to be created for each '
                                                   f'period. This will overwrite any files already created. Are you sure '
                                                   f'you want to do this?')
        if answer == 'yes':
            try:
                logging.debug('Creating new excel files from text files.')
                for file in os.listdir('.'):
                    if not file.lower().endswith('.txt'):
                        continue
                    period = file.replace('.txt', '')
                    wb = Workbook()
                    names = open(file).readlines()
                    for name in names:
                        write_excel_page(wb, name)
                    wb.remove(wb['Sheet'])
                    wb.save(period + '.xlsx')
                tk.messagebox.showinfo(title='Update Message',
                                       message='Files created successfully. Please restart the program.')
                self.refresh_ui()
                logging.debug('Excel files created.')
            except:
                logging.debug('Error occurred when trying to write excel files.')


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    set_icon(root)
    app.mainloop()
